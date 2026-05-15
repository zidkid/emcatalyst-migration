"""
Bulk BRS Upload — accepts an Excel file and creates multiple BRS applications with doctors.

Excel format: Each row is one doctor. Rows with the same BRS title + survey_title get grouped into one BRS.

Key changes from manual BRS creation:
- Uses survey_title (not survey_id) to look up the survey
- Uses division_name (not division_id) to look up the division
- Uses doctor_uid to look up doctors from HcpDoctor table by uid_number
- If doctor_uid not found in DB, that row is skipped with an error
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime
from decimal import Decimal, InvalidOperation
import io

from app.db.base import get_db
from app.api.deps import get_current_user
from app.models.user import User, Division
from app.models.brs import BrsApplication, BrsDoctor, BrsStatus, BrsSurvey, BrsAuditTrail
from app.models.master import HcpDoctor

router = APIRouter(prefix="/brs/bulk", tags=["BRS Bulk"])


BRS_GROUP_FIELDS = [
    "survey_title", "title", "division_name", "therapeutic_area", "brand",
    "budget_type", "platform", "topic", "on_field_execution_by",
    "start_date", "end_date", "city", "venue", "rationale", "agenda",
    "cost_center", "remarks",
]

DOCTOR_FIELDS = ["doctor_uid", "name_as_per_pan", "pan_number", "email", "honorarium_amount"]

ALL_COLUMNS = BRS_GROUP_FIELDS + DOCTOR_FIELDS


def _generate_brs_code(db: Session) -> str:
    count = db.query(BrsApplication).count() + 1
    return f"BRS{datetime.now().strftime('%y%m')}{count:04d}"


def _has_role(user: User, role_name: str) -> bool:
    if user.role == role_name:
        return True
    if user.role == "Administrator" or user.is_superuser:
        return True
    if user.role_assignments:
        for ra in user.role_assignments:
            if ra.role == role_name:
                return True
    return False


@router.get("/template")
def download_template():
    """Download an Excel template for bulk BRS upload."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BRS Bulk Upload"

    headers = [
        "survey_title", "title", "division_name", "therapeutic_area", "brand",
        "budget_type", "platform", "topic", "on_field_execution_by",
        "start_date", "end_date", "city", "venue", "rationale", "agenda",
        "cost_center", "remarks",
        "doctor_uid", "name_as_per_pan", "pan_number", "email", "honorarium_amount",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Example rows (2 doctors for same BRS)
    example1 = [
        "Survey Template Name", "BRS Program Title", "Division Name", "Cardiology", "BrandX",
        "Head Office", "In Person", "Topic here", "EMP001",
        "2026-06-01", "2026-06-15", "Mumbai", "Hotel XYZ", "Rationale text", "Agenda text",
        "CC001", "Optional remarks",
        "UID001", "Dr John Doe", "ABCDE1234F", "john@example.com", 25000,
    ]
    example2 = [
        "Survey Template Name", "BRS Program Title", "Division Name", "", "",
        "", "", "", "",
        "", "", "", "", "", "",
        "", "",
        "UID002", "Dr Jane Smith", "FGHIJ5678K", "jane@example.com", 30000,
    ]
    for col, val in enumerate(example1, 1):
        ws.cell(row=2, column=col, value=val)
    for col, val in enumerate(example2, 1):
        ws.cell(row=3, column=col, value=val)

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        "BULK BRS UPLOAD INSTRUCTIONS",
        "",
        "1. Each row represents ONE DOCTOR in a BRS application.",
        "2. Multiple doctors for the same BRS: repeat 'title' + 'survey_title' on each row.",
        "   (Other BRS fields can be left blank on subsequent rows — they are read from the first row.)",
        "3. Rows are grouped by 'title' + 'survey_title' — same combination = same BRS.",
        "",
        "REQUIRED FIELDS:",
        "- survey_title: Must match an existing survey name in the system (exact match)",
        "- title: BRS program title (also used for grouping)",
        "- doctor_uid: UID number of the doctor from the MCL (Master Contact List)",
        "  → If the UID is not found in the system, that doctor row will be SKIPPED.",
        "",
        "DOCTOR FIELDS (override MCL defaults per survey):",
        "- doctor_uid: Matched against HcpDoctor.uid_number (REQUIRED)",
        "- name_as_per_pan: Name as per PAN card (if blank, uses MCL value)",
        "- pan_number: PAN card number (if blank, uses MCL value)",
        "- email: Email address (if blank, uses MCL value)",
        "- honorarium_amount: Amount in INR (numeric, no currency symbol)",
        "",
        "NOTE: name_as_per_pan, pan_number, and email can vary per survey for the same doctor.",
        "If provided in the Excel, they override the MCL defaults for this BRS entry.",
        "If left blank, the values are auto-filled from the MCL record.",
        "",
        "LOOKUP FIELDS (matched by name, not ID):",
        "- survey_title: Matched against BrsSurvey.title",
        "- division_name: Matched against Division.name (leave blank to use your default division)",
        "",
        "OTHER FIELDS:",
        "- therapeutic_area: e.g. Cardiology, Oncology",
        "- brand: Product brand name",
        "- budget_type: Head Office / Field",
        "- platform: In Person / Virtual / Both",
        "- topic: Topic/subject of the program",
        "- on_field_execution_by: Employee ID of the Territory Manager who will execute on field",
        "- start_date / end_date: YYYY-MM-DD format",
        "- city / venue: Location details",
        "- rationale: Business rationale",
        "- agenda: Program agenda",
        "- cost_center: Cost center code",
        "- remarks: Additional remarks",
    ]
    for i, line in enumerate(instructions, 1):
        ws2.cell(row=i, column=1, value=line)
    ws2.column_dimensions['A'].width = 90

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=brs_bulk_template.xlsx"},
    )


@router.post("/upload")
async def upload_bulk_brs(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload an Excel file to create multiple BRS applications with doctors.
    - survey_title is looked up by name
    - division_name is looked up by name
    - doctor_uid is looked up from HcpDoctor.uid_number — if not found, row is skipped
    """
    # Check RBAC: user must have access to brs_bulk_upload page
    from app.services.rbac_service import get_user_accessible_pages
    accessible = get_user_accessible_pages(db, current_user)
    if "brs_bulk_upload" not in accessible:
        raise HTTPException(403, "You do not have access to bulk upload")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only .xlsx or .xls files are accepted")

    import openpyxl

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Failed to read Excel file: {str(e)}")

    # Read headers
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip().lower() if cell.value else "")

    if "title" not in headers or "doctor_uid" not in headers:
        raise HTTPException(400, "Excel must have at least 'title' and 'doctor_uid' columns")
    if "survey_title" not in headers:
        raise HTTPException(400, "Excel must have a 'survey_title' column")

    # Parse rows
    rows = []
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        row_data = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                row_data[headers[col_idx]] = value
        if not row_data.get("title"):
            errors.append({"row": row_idx, "error": "Missing 'title'"})
            continue
        if not row_data.get("doctor_uid"):
            errors.append({"row": row_idx, "error": "Missing 'doctor_uid'"})
            continue
        rows.append({"row_idx": row_idx, "data": row_data})

    if not rows:
        raise HTTPException(400, "No valid data rows found in the file")

    # Pre-fetch lookups for performance
    # Surveys by title
    all_surveys = db.query(BrsSurvey).filter(BrsSurvey.is_active == True).all()
    survey_map = {s.title.strip().lower(): s for s in all_surveys}

    # Divisions by name
    all_divisions = db.query(Division).filter(Division.is_active == True).all()
    division_map = {d.name.strip().lower(): d for d in all_divisions}

    # Doctors by uid_number
    uid_list = [str(r["data"].get("doctor_uid", "")).strip() for r in rows if r["data"].get("doctor_uid")]
    doctors_in_db = db.query(HcpDoctor).filter(HcpDoctor.uid_number.in_(uid_list)).all() if uid_list else []
    doctor_map = {d.uid_number.strip(): d for d in doctors_in_db if d.uid_number}

    # Territory Managers by employee_id (for on_field_execution_by validation)
    from app.models.user import UserRoleAssignment
    tm_users = (
        db.query(User)
        .join(UserRoleAssignment, UserRoleAssignment.user_id == User.id)
        .filter(UserRoleAssignment.role == "Territory Manager", User.is_active == True)
        .all()
    )
    tm_map = {u.employee_id: u for u in tm_users if u.employee_id}

    # Pre-fetch survey-doctor mappings for validation
    from app.models.brs import SurveyDoctorMapping
    all_survey_mappings = db.query(SurveyDoctorMapping).all()
    # Build a set of (survey_id, hcp_doctor_id) for fast lookup
    survey_doctor_set = {(m.survey_id, m.hcp_doctor_id) for m in all_survey_mappings}

    # Group rows by BRS (title + survey_title)
    brs_groups = {}
    for item in rows:
        data = item["data"]
        group_key = f"{str(data.get('survey_title', '')).strip().lower()}|{str(data.get('title', '')).strip().lower()}"

        if group_key not in brs_groups:
            brs_groups[group_key] = {"brs_data": {}, "doctors": []}
            for field in BRS_GROUP_FIELDS:
                val = data.get(field)
                if val is not None and str(val).strip():
                    brs_groups[group_key]["brs_data"][field] = val

        # Doctor lookup
        uid = str(data.get("doctor_uid", "")).strip()
        doctor = doctor_map.get(uid)
        if not doctor:
            errors.append({"row": item["row_idx"], "error": f"Doctor UID '{uid}' not found in MCL"})
            continue

        brs_groups[group_key]["doctors"].append({
            "hcp_doctor": doctor,
            "honorarium_amount": data.get("honorarium_amount"),
            "name_as_per_pan": data.get("name_as_per_pan"),
            "pan_number": data.get("pan_number"),
            "email": data.get("email"),
            "_row": item["row_idx"],
        })

    # Create BRS applications
    created = []
    for group_key, group in brs_groups.items():
        brs_data = group["brs_data"]
        doctors = group["doctors"]

        if not doctors:
            continue  # All doctors in this group were invalid

        # Resolve survey
        survey_title = str(brs_data.get("survey_title", "")).strip().lower()
        survey = survey_map.get(survey_title)
        if not survey:
            errors.append({"row": doctors[0]["_row"], "error": f"Survey '{brs_data.get('survey_title')}' not found"})
            continue

        # Validate doctors are mapped to this survey (if survey has mappings)
        survey_has_mappings = any(s_id == survey.id for s_id, _ in survey_doctor_set)
        if survey_has_mappings:
            valid_doctors = []
            for doc_entry in doctors:
                hcp = doc_entry["hcp_doctor"]
                if (survey.id, hcp.id) not in survey_doctor_set:
                    errors.append({"row": doc_entry["_row"], "error": f"Doctor '{hcp.full_name or hcp.uid_number}' is not mapped to survey '{survey.title}'"})
                else:
                    valid_doctors.append(doc_entry)
            doctors = valid_doctors
            if not doctors:
                continue

        # Resolve division
        division_id = current_user.division_id
        division_name = str(brs_data.get("division_name", "")).strip().lower()
        if division_name:
            div = division_map.get(division_name)
            if div:
                division_id = div.id
            else:
                errors.append({"row": doctors[0]["_row"], "error": f"Division '{brs_data.get('division_name')}' not found, using default"})

        # Parse dates
        start_date = _parse_date(brs_data.get("start_date"))
        end_date = _parse_date(brs_data.get("end_date"))

        # Resolve on_field_execution_by (employee_id of Territory Manager)
        on_field_value = _str_or_none(brs_data.get("on_field_execution_by"))
        if on_field_value:
            tm = tm_map.get(on_field_value.strip())
            if tm:
                on_field_value = tm.employee_id  # Store employee_id
            else:
                errors.append({"row": doctors[0]["_row"], "error": f"Territory Manager with employee_id '{on_field_value}' not found"})
                on_field_value = None

        # Create BRS
        app = BrsApplication(
            brs_code=_generate_brs_code(db),
            survey_id=survey.id,
            division_id=division_id,
            title=str(brs_data.get("title", "")),
            remarks=_str_or_none(brs_data.get("remarks")),
            therapeutic_area=_str_or_none(brs_data.get("therapeutic_area")),
            brand=_str_or_none(brs_data.get("brand")),
            budget_type=_str_or_none(brs_data.get("budget_type")),
            platform=_str_or_none(brs_data.get("platform")),
            topic=_str_or_none(brs_data.get("topic")),
            on_field_execution_by=on_field_value,
            start_date=start_date,
            end_date=end_date,
            city=_str_or_none(brs_data.get("city")),
            venue=_str_or_none(brs_data.get("venue")),
            rationale=_str_or_none(brs_data.get("rationale")),
            agenda=_str_or_none(brs_data.get("agenda")),
            cost_center=_str_or_none(brs_data.get("cost_center")),
            status=BrsStatus.DRAFT,
            created_by_id=current_user.id,
        )
        db.add(app)
        db.flush()

        # Audit trail
        db.add(BrsAuditTrail(
            application_id=app.id,
            action="Bulk Created",
            from_status="",
            to_status=BrsStatus.DRAFT,
            performed_by_id=current_user.id,
            remarks=f"Bulk upload: {len(doctors)} doctors",
        ))

        # Add doctors from MCL
        doctor_count = 0
        for doc_entry in doctors:
            hcp = doc_entry["hcp_doctor"]
            honorarium = None
            if doc_entry.get("honorarium_amount"):
                try:
                    honorarium = Decimal(str(doc_entry["honorarium_amount"]))
                except (InvalidOperation, ValueError):
                    errors.append({"row": doc_entry["_row"], "error": f"Invalid honorarium_amount: {doc_entry['honorarium_amount']}"})
                    continue

            doctor = BrsDoctor(
                brs_application_id=app.id,
                hcp_doctor_id=hcp.id,
                doctor_name=hcp.full_name or f"{hcp.first_name or ''} {hcp.last_name or ''}".strip(),
                name_as_per_pan=_str_or_none(doc_entry.get("name_as_per_pan")) or hcp.full_name or f"{hcp.first_name or ''} {hcp.last_name or ''}".strip(),
                pan_number=_str_or_none(doc_entry.get("pan_number")) or hcp.pan_number,
                email=_str_or_none(doc_entry.get("email")) or hcp.email,
                mobile=hcp.mobile_number,
                speciality=hcp.speciality,
                honorarium_amount=honorarium,
            )
            db.add(doctor)
            doctor_count += 1

        created.append({
            "brs_code": app.brs_code,
            "title": app.title,
            "doctor_count": doctor_count,
            "id": app.id,
        })

    db.commit()

    return {
        "message": f"Successfully created {len(created)} BRS application(s)",
        "created": created,
        "errors": errors,
        "total_rows_processed": len(rows),
        "doctors_skipped": len([e for e in errors if "not found in MCL" in e.get("error", "")]),
    }


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d")
    except ValueError:
        try:
            return datetime.strptime(str(value).strip(), "%d-%m-%Y")
        except ValueError:
            return None


def _str_or_none(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None



# ─── Survey Bulk Import ───────────────────────────────────────────────────────

@router.get("/survey-template")
def download_survey_template():
    """Download an Excel template for bulk survey creation with questions."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Survey Import"

    headers = [
        "survey_title", "survey_description", "division_name",
        "total_honorarium_amount", "question_text", "question_type",
        "options", "is_required",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    # Example rows showing all question types
    examples = [
        # Row 2: Free text question
        [
            "Cardiology Expert Survey", "Survey for cardiology experts", "Oncology",
            50000, "What is your experience with drug X in clinical practice?",
            "free_text", "", "Yes",
        ],
        # Row 3: Single select question
        [
            "Cardiology Expert Survey", "", "",
            "", "How often do you prescribe drug X?",
            "single_select", "Daily|Weekly|Monthly|Rarely|Never", "Yes",
        ],
        # Row 4: Multi select question
        [
            "Cardiology Expert Survey", "", "",
            "", "Which conditions do you treat with drug X? (select all that apply)",
            "multi_select", "Hypertension|Heart Failure|Arrhythmia|Post-MI|Other", "Yes",
        ],
        # Row 5: Fill in the blanks question
        [
            "Cardiology Expert Survey", "", "",
            "", "I prescribe drug X for ___ patients per month and find it most effective for ___ condition.",
            "fill_in_blanks", "", "Yes",
        ],
        # Row 6: Another free text (optional)
        [
            "Cardiology Expert Survey", "", "",
            "", "Any additional comments or suggestions?",
            "free_text", "", "No",
        ],
    ]
    for row_idx, row_data in enumerate(examples, 2):
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=val)

    # Doctors sheet — for mapping doctors to the survey
    ws_docs = wb.create_sheet("Doctors")
    ws_docs.cell(row=1, column=1, value="survey_title").font = openpyxl.styles.Font(bold=True)
    ws_docs.cell(row=1, column=2, value="doctor_uid").font = openpyxl.styles.Font(bold=True)
    ws_docs.column_dimensions['A'].width = 30
    ws_docs.column_dimensions['B'].width = 20
    # Example doctor rows
    ws_docs.cell(row=2, column=1, value="Cardiology Expert Survey")
    ws_docs.cell(row=2, column=2, value="UID001")
    ws_docs.cell(row=3, column=1, value="Cardiology Expert Survey")
    ws_docs.cell(row=3, column=2, value="UID002")
    ws_docs.cell(row=4, column=1, value="Cardiology Expert Survey")
    ws_docs.cell(row=4, column=2, value="UID003")

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        "SURVEY BULK IMPORT INSTRUCTIONS",
        "",
        "This file has TWO sheets:",
        "  1. 'Survey Import' — questions for the survey",
        "  2. 'Doctors' — doctor UIDs to map to the survey (MANDATORY)",
        "",
        "SHEET 1: SURVEY IMPORT (Questions)",
        "1. Each row represents ONE QUESTION in a survey.",
        "2. All rows with the same 'survey_title' are grouped into one survey.",
        "3. Survey metadata (description, division, honorarium) is read from the first row of each group.",
        "4. Questions are added in the order they appear in the file.",
        "",
        "REQUIRED FIELDS:",
        "- survey_title: Name of the survey (used for grouping)",
        "- question_text: The question to ask the doctor",
        "- question_type: One of the following types:",
        "",
        "QUESTION TYPES:",
        "- free_text: Open-ended text answer (doctor types freely)",
        "- single_select: Doctor picks ONE option from a list",
        "- multi_select: Doctor picks MULTIPLE options from a list",
        "- fill_in_blanks: Question with blanks (___) that the doctor fills in",
        "",
        "OPTIONS FIELD:",
        "- For single_select and multi_select: provide options separated by | (pipe)",
        "- For free_text and fill_in_blanks: leave blank",
        "",
        "IS_REQUIRED: Yes or No (default: Yes)",
        "",
        "SHEET 2: DOCTORS (Mandatory)",
        "- Column A: survey_title (must match the survey title in Sheet 1)",
        "- Column B: doctor_uid (UID number from MCL)",
        "- At least ONE doctor must be mapped per survey.",
        "- Only mapped doctors will be available when creating BRS with this survey.",
    ]
    for i, line in enumerate(instructions, 1):
        ws2.cell(row=i, column=1, value=line)
    ws2.column_dimensions['A'].width = 90

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=survey_import_template.xlsx"},
    )


@router.post("/survey-upload")
async def upload_bulk_survey(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Upload an Excel file to create surveys with questions and doctor mappings.
    Sheet 1: Questions (grouped by survey_title)
    Sheet 2 (Doctors): doctor_uid mappings per survey (MANDATORY)
    """
    from app.models.brs import BrsSurveyQuestion, BrsQuestionType, SurveyDoctorMapping

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only .xlsx or .xls files are accepted")

    import openpyxl

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        # Find the questions sheet
        ws = None
        for sheet_name in wb.sheetnames:
            if 'survey' in sheet_name.lower():
                ws = wb[sheet_name]
                break
        if ws is None:
            for sheet_name in wb.sheetnames:
                if sheet_name.lower() not in ('instructions', 'doctors'):
                    ws = wb[sheet_name]
                    break
        if ws is None:
            ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Failed to read Excel file: {str(e)}")

    # Read "Doctors" sheet (mandatory)
    doctor_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == 'doctors':
            doctor_sheet = wb[sheet_name]
            break

    # Parse doctor UIDs from Doctors sheet
    doctor_uids_by_survey = {}  # {survey_title_lower: [uid1, uid2, ...]}
    if doctor_sheet:
        doc_headers = []
        for cell in doctor_sheet[1]:
            doc_headers.append(str(cell.value).strip().lower() if cell.value else "")
        # Detect offset
        doc_offset = 0
        for h in doc_headers:
            if h == "":
                doc_offset += 1
            else:
                break
        doc_headers = doc_headers[doc_offset:]
        for row in doctor_sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            shifted = row[doc_offset:] if doc_offset else row
            row_data = {}
            for col_idx, value in enumerate(shifted):
                if col_idx < len(doc_headers) and doc_headers[col_idx]:
                    row_data[doc_headers[col_idx]] = value
            s_title = str(row_data.get("survey_title", "")).strip().lower()
            uid = str(row_data.get("doctor_uid", "")).strip()
            if s_title and uid:
                if s_title not in doctor_uids_by_survey:
                    doctor_uids_by_survey[s_title] = []
                doctor_uids_by_survey[s_title].append(uid)

    # Read headers from questions sheet (detect and skip empty leading columns)
    raw_headers = []
    for cell in ws[1]:
        raw_headers.append(str(cell.value).strip().lower() if cell.value else "")

    # Detect column offset (empty leading columns)
    col_offset = 0
    for h in raw_headers:
        if h == "":
            col_offset += 1
        else:
            break
    headers = raw_headers[col_offset:]

    if "survey_title" not in headers or "question_text" not in headers:
        raise HTTPException(400, "Excel must have 'survey_title' and 'question_text' columns")

    # Valid question types
    valid_types = {t.value for t in BrsQuestionType}

    # Parse rows
    rows = []
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        # Apply same column offset
        shifted_row = row[col_offset:] if col_offset else row
        row_data = {}
        for col_idx, value in enumerate(shifted_row):
            if col_idx < len(headers) and headers[col_idx]:
                row_data[headers[col_idx]] = value
        if not row_data.get("survey_title"):
            errors.append({"row": row_idx, "error": "Missing 'survey_title'"})
            continue
        if not row_data.get("question_text"):
            errors.append({"row": row_idx, "error": "Missing 'question_text'"})
            continue
        # Validate question type
        q_type = str(row_data.get("question_type", "free_text")).strip().lower()
        if q_type not in valid_types:
            errors.append({"row": row_idx, "error": f"Invalid question_type '{q_type}'. Must be one of: {', '.join(valid_types)}"})
            continue
        row_data["question_type"] = q_type
        rows.append({"row_idx": row_idx, "data": row_data})

    if not rows:
        raise HTTPException(400, "No valid data rows found in the file")

    # Division lookup
    all_divisions = db.query(Division).filter(Division.is_active == True).all()
    division_map = {d.name.strip().lower(): d for d in all_divisions}

    # Group by survey_title
    survey_groups = {}
    for item in rows:
        data = item["data"]
        key = str(data["survey_title"]).strip().lower()
        if key not in survey_groups:
            survey_groups[key] = {"meta": data, "questions": []}
        survey_groups[key]["questions"].append({
            "question_text": str(data["question_text"]),
            "question_type": data["question_type"],
            "options": str(data.get("options", "")).split("|") if data.get("options") else [],
            "is_required": str(data.get("is_required", "Yes")).strip().lower() in ("yes", "true", "1", ""),
            "_row": item["row_idx"],
        })

    # Create surveys
    created = []
    for key, group in survey_groups.items():
        meta = group["meta"]
        questions = group["questions"]

        # Check if survey with same title already exists
        existing = db.query(BrsSurvey).filter(BrsSurvey.title == str(meta["survey_title"]).strip()).first()
        if existing:
            errors.append({"row": questions[0]["_row"], "error": f"Survey '{meta['survey_title']}' already exists (ID: {existing.id})"})
            continue

        # Resolve division
        division_id = None
        div_name = str(meta.get("division_name", "")).strip().lower()
        if div_name:
            div = division_map.get(div_name)
            if div:
                division_id = div.id
            else:
                errors.append({"row": questions[0]["_row"], "error": f"Division '{meta.get('division_name')}' not found, skipping division assignment"})

        # Parse honorarium
        honorarium = None
        if meta.get("total_honorarium_amount"):
            try:
                honorarium = Decimal(str(meta["total_honorarium_amount"]))
            except (InvalidOperation, ValueError):
                pass

        # Create survey
        survey = BrsSurvey(
            title=str(meta["survey_title"]).strip(),
            description=_str_or_none(meta.get("survey_description")),
            division_id=division_id,
            total_honorarium_amount=honorarium,
            is_active=True,
            created_by_id=current_user.id,
        )
        db.add(survey)
        db.flush()

        # Add questions
        for order, q_data in enumerate(questions, 1):
            options = [o.strip() for o in q_data["options"] if o.strip()]
            question = BrsSurveyQuestion(
                survey_id=survey.id,
                order_no=order,
                question_text=q_data["question_text"],
                question_type=q_data["question_type"],
                options=options if options else None,
                is_required=q_data["is_required"],
            )
            db.add(question)

        # Map doctors to survey (mandatory)
        survey_key = str(meta["survey_title"]).strip().lower()
        uids_for_survey = doctor_uids_by_survey.get(survey_key, [])
        if not uids_for_survey:
            errors.append({"row": questions[0]["_row"], "error": f"No doctors found in 'Doctors' sheet for survey '{meta['survey_title']}'. At least one doctor is required."})
            db.rollback()
            continue

        doctors_mapped = 0
        doctors_not_found = []
        for uid in uids_for_survey:
            doc = db.query(HcpDoctor).filter(HcpDoctor.uid_number == uid).first()
            if not doc:
                doctors_not_found.append(uid)
                continue
            db.add(SurveyDoctorMapping(survey_id=survey.id, hcp_doctor_id=doc.id))
            doctors_mapped += 1

        if doctors_mapped == 0:
            errors.append({"row": questions[0]["_row"], "error": f"None of the doctor UIDs in 'Doctors' sheet were found in MCL for survey '{meta['survey_title']}'"})
            db.rollback()
            continue

        if doctors_not_found:
            errors.append({"row": questions[0]["_row"], "error": f"Some doctor UIDs not found: {', '.join(doctors_not_found[:5])}"})

        created.append({
            "id": survey.id,
            "title": survey.title,
            "question_count": len(questions),
            "doctors_mapped": doctors_mapped,
        })

    db.commit()

    return {
        "message": f"Successfully created {len(created)} survey(s)",
        "created": created,
        "errors": errors,
        "total_rows_processed": len(rows),
    }



@router.post("/survey-doctors-import")
async def import_survey_doctors_excel(
    survey_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import doctor UIDs from an Excel/CSV file and map them to a survey."""
    from app.models.brs import SurveyDoctorMapping, BrsSurvey

    survey = db.query(BrsSurvey).filter(BrsSurvey.id == survey_id).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    uids = []

    if file.filename.endswith(('.xlsx', '.xls')):
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, values_only=True):
            for cell in row:
                if cell and str(cell).strip() and str(cell).strip().lower() != 'doctor_uid':
                    uids.append(str(cell).strip())
    else:
        content = await file.read()
        text = content.decode('utf-8', errors='ignore')
        uids = [s.strip() for s in text.split('\n') if s.strip() and s.strip().lower() != 'doctor_uid']

    if not uids:
        raise HTTPException(400, "No UIDs found in file")

    added = 0
    not_found = []
    for uid in uids:
        doc = db.query(HcpDoctor).filter(HcpDoctor.uid_number == uid).first()
        if not doc:
            not_found.append(uid)
            continue
        existing = db.query(SurveyDoctorMapping).filter(
            SurveyDoctorMapping.survey_id == survey_id,
            SurveyDoctorMapping.hcp_doctor_id == doc.id
        ).first()
        if not existing:
            db.add(SurveyDoctorMapping(survey_id=survey_id, hcp_doctor_id=doc.id))
            added += 1
    db.commit()
    return {"added": added, "not_found": not_found, "total": db.query(SurveyDoctorMapping).filter(SurveyDoctorMapping.survey_id == survey_id).count()}
